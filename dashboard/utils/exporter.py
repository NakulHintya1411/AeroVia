"""Export route analysis to Excel and PDF."""
import io
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Excel Export ──────────────────────────────────────────────────────────────
def export_excel(econ, scenarios=None) -> bytes:
    """Generate Excel workbook with route analysis + scenario comparison."""
    wb = openpyxl.Workbook()

    # Colours
    BLUE   = "2563EB"; LBLUE  = "EFF4FF"; GREEN = "059669"
    LGREEN = "ECFDF5"; RED    = "DC2626"; LRED  = "FEF2F2"
    GREY   = "F1F4F9"; DGREY  = "3D4F6B"; WHITE = "FFFFFF"

    def hdr_style(ws, row, col, val, bg=BLUE, fg=WHITE, bold=True):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, color=fg, name="Calibri", size=10)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="D0D7E6")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return c

    def val_style(ws, row, col, val, fmt=None, bg=WHITE, bold=False, align="right"):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=10, bold=bold, color=DGREY)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        thin = Side(style="thin", color="E8EDF5")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        if fmt: c.number_format = fmt
        return c

    # ── Sheet 1: Route Summary ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Route Analysis"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.row_dimensions[1].height = 30

    # Title
    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = f"AeroVia Route Analysis — {econ.origin} → {econ.destination}"
    t.font = Font(bold=True, size=14, color=BLUE, name="Calibri")
    t.fill = PatternFill("solid", fgColor=LBLUE)
    t.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:B2")
    d = ws["A2"]
    d.value = f"Generated: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}  |  Airline: {econ.airline_code}  |  Aircraft: {econ.aircraft_type}"
    d.font = Font(size=9, color="7A8BA8", name="Calibri")
    d.fill = PatternFill("solid", fgColor=GREY)

    # Section headers + data
    sections = [
        ("ROUTE DETAILS", [
            ("Origin", econ.origin),
            ("Destination", econ.destination),
            ("Distance", f"{econ.distance_km:,.0f} km"),
            ("Flight Time", f"{econ.flight_time_hr:.1f} hr"),
            ("Aircraft", f"{econ.aircraft_type} ({econ.seats} seats)"),
            ("Daily Frequency", f"{econ.daily_frequency} flights"),
            ("Load Factor", f"{econ.load_factor:.1%}"),
        ]),
        ("KEY METRICS", [
            ("RASK", f"{econ.rask:.2f} paise/ASK"),
            ("CASK", f"{econ.cask:.2f} paise/ASK"),
            ("Net Margin", f"{econ.margin_pct:.1f}%"),
            ("Break-Even LF", f"{econ.belf:.1%}"),
            ("LF Cushion", f"{econ.lf_cushion:+.1%}"),
            ("Daily ASK", f"{econ.ask:,.0f}"),
            ("Daily RPK", f"{econ.rpk:,.0f}"),
        ]),
        ("DAILY P&L (₹)", [
            ("Total Revenue", econ.total_revenue_inr),
            ("Fuel Cost", econ.fuel_cost_inr),
            ("Crew Cost", econ.crew_cost_inr),
            ("Maintenance", econ.maintenance_cost_inr),
            ("Airport Charges", econ.airport_cost_inr),
            ("Overhead", econ.overhead_cost_inr),
            ("Total Cost", econ.total_cost_inr),
            ("Net Profit / Loss", econ.profit_inr),
        ]),
    ]

    row = 4
    for sec_title, rows in sections:
        ws.merge_cells(f"A{row}:B{row}")
        hdr_style(ws, row, 1, sec_title, bg=BLUE, fg=WHITE)
        ws.row_dimensions[row].height = 22
        row += 1
        for label, value in rows:
            bg = LGREEN if (label == "Net Profit / Loss" and isinstance(value, (int,float)) and value >= 0) else \
                 LRED   if (label == "Net Profit / Loss" and isinstance(value, (int,float)) and value < 0) else WHITE
            val_style(ws, row, 1, label, align="left", bg=bg, bold=True)
            fmt = '#,##0.00' if isinstance(value, float) and "%" not in str(value) and "p" not in str(value) else None
            val_style(ws, row, 2, value, bg=bg, fmt=fmt)
            ws.row_dimensions[row].height = 18
            row += 1
        row += 1

    # ── Sheet 2: Scenario Comparison ──────────────────────────────────────
    if scenarios:
        ws2 = wb.create_sheet("Scenario Comparison")
        ws2.sheet_view.showGridLines = False

        cols = ["Scenario", "Route", "Aircraft", "LF %", "RASK (p)",
                "CASK (p)", "Margin %", "BELF %", "Daily Profit (₹)"]
        for i, col in enumerate(cols, 1):
            ws2.column_dimensions[get_column_letter(i)].width = 18
            hdr_style(ws2, 1, i, col)

        for r, sc in enumerate(scenarios, 2):
            e = sc["economics"]
            profit = e["profit_inr"]
            row_bg = LGREEN if profit >= 0 else LRED
            data = [sc["name"], f"{e['origin']}→{e['destination']}",
                    e["aircraft_type"], f"{e['load_factor']*100:.0f}%",
                    f"{e['rask']:.1f}", f"{e['cask']:.1f}",
                    f"{e['margin_pct']:.1f}%", f"{e['belf']*100:.1f}%",
                    f"₹{profit/1e5:.1f}L"]
            for i, val in enumerate(data, 1):
                val_style(ws2, r, i, val, bg=row_bg if i == len(data) else WHITE)
            ws2.row_dimensions[r].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── PDF Export ────────────────────────────────────────────────────────────────
def export_pdf(econ, scenarios=None) -> bytes:
    """Generate PDF route analysis report."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=18*mm, rightMargin=18*mm,
                             topMargin=18*mm, bottomMargin=18*mm)

    BLUE   = colors.HexColor("#2563EB")
    LBLUE  = colors.HexColor("#EFF4FF")
    GREEN  = colors.HexColor("#059669")
    LGREEN = colors.HexColor("#ECFDF5")
    RED    = colors.HexColor("#DC2626")
    LRED   = colors.HexColor("#FEF2F2")
    GREY   = colors.HexColor("#F1F4F9")
    DGREY  = colors.HexColor("#3D4F6B")
    LGREY  = colors.HexColor("#E2E7F0")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Normal"],
        fontSize=18, fontName="Helvetica-Bold", textColor=BLUE,
        spaceAfter=2*mm, leading=22)
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"],
        fontSize=9, fontName="Helvetica", textColor=DGREY, spaceAfter=4*mm)
    sec_style = ParagraphStyle("Sec", parent=styles["Normal"],
        fontSize=10, fontName="Helvetica-Bold", textColor=colors.white,
        spaceAfter=0, leading=14)
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"],
        fontSize=9, fontName="Helvetica", textColor=DGREY)

    def section_table(title, rows, col_widths):
        data = [[Paragraph(f"  {title}", sec_style), ""]]
        for label, value in rows:
            data.append([label, str(value)])
        style = TableStyle([
            ("BACKGROUND", (0,0), (-1,0), BLUE),
            ("SPAN",       (0,0), (-1,0)),
            ("FONTNAME",   (0,1), (-1,-1), "Helvetica"),
            ("FONTSIZE",   (0,1), (-1,-1), 9),
            ("TEXTCOLOR",  (0,1), (-1,-1), DGREY),
            ("FONTNAME",   (0,1), (0,-1), "Helvetica-Bold"),
            ("BACKGROUND", (0,1), (-1,-1), colors.white),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, GREY]),
            ("GRID",       (0,0), (-1,-1), 0.5, LGREY),
            ("LEFTPADDING",  (0,0), (-1,-1), 8),
            ("RIGHTPADDING", (0,0), (-1,-1), 8),
            ("TOPPADDING",   (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0), (-1,-1), 5),
            ("ROWHEIGHT",  (0,0), (0,0), 18),
        ])
        # Highlight profit/loss row
        for i, (label, value) in enumerate(rows, 1):
            if "Profit" in label or "Loss" in label:
                bg = LGREEN if isinstance(value, (int,float)) and value >= 0 else LRED
                style.add("BACKGROUND", (0,i), (-1,i), bg)
        return Table(data, colWidths=col_widths, style=style, hAlign="LEFT")

    story = []

    # Header
    story.append(Paragraph(f"AeroVia Route Analysis", title_style))
    story.append(Paragraph(
        f"{econ.origin} → {econ.destination}  |  {econ.airline_code}  |  "
        f"{econ.aircraft_type}  |  Generated {datetime.datetime.now().strftime('%d %b %Y')}",
        sub_style))
    story.append(HRFlowable(width="100%", thickness=1, color=LBLUE, spaceAfter=4*mm))

    W = 174*mm
    half = W/2 - 3*mm

    # Route details + Key metrics side by side
    left = section_table("ROUTE DETAILS", [
        ("Origin Airport",   econ.origin),
        ("Destination",      econ.destination),
        ("Distance",         f"{econ.distance_km:,.0f} km"),
        ("Flight Time",      f"{econ.flight_time_hr:.1f} hr"),
        ("Aircraft",         f"{econ.aircraft_type} ({econ.seats} seats)"),
        ("Daily Frequency",  f"{econ.daily_frequency} flights"),
        ("Load Factor",      f"{econ.load_factor:.1%}"),
    ], [half*0.55, half*0.45])

    right = section_table("KEY METRICS", [
        ("RASK",             f"{econ.rask:.2f}p/ASK"),
        ("CASK",             f"{econ.cask:.2f}p/ASK"),
        ("Net Margin",       f"{econ.margin_pct:.1f}%"),
        ("Break-Even LF",    f"{econ.belf:.1%}"),
        ("LF Cushion",       f"{econ.lf_cushion:+.1%}"),
        ("Daily ASK",        f"{econ.ask:,.0f}"),
        ("Daily RPK",        f"{econ.rpk:,.0f}"),
    ], [half*0.55, half*0.45])

    combo = Table([[left, right]], colWidths=[half, half],
                   style=TableStyle([("LEFTPADDING",(0,0),(-1,-1),0),
                                      ("RIGHTPADDING",(0,0),(-1,-1),6),
                                      ("VALIGN",(0,0),(-1,-1),"TOP")]))
    story.append(combo)
    story.append(Spacer(1, 5*mm))

    # P&L table
    story.append(section_table("DAILY PROFIT & LOSS  (₹)", [
        ("Total Revenue",    f"₹{econ.total_revenue_inr:,.0f}"),
        ("Fuel Cost",        f"₹{econ.fuel_cost_inr:,.0f}"),
        ("Crew Cost",        f"₹{econ.crew_cost_inr:,.0f}"),
        ("Maintenance",      f"₹{econ.maintenance_cost_inr:,.0f}"),
        ("Airport Charges",  f"₹{econ.airport_cost_inr:,.0f}"),
        ("Overhead",         f"₹{econ.overhead_cost_inr:,.0f}"),
        ("Total Cost",       f"₹{econ.total_cost_inr:,.0f}"),
        ("Net Profit / Loss",econ.profit_inr),
    ], [W*0.5, W*0.5]))
    story.append(Spacer(1, 5*mm))

    # Scenario comparison
    if scenarios:
        story.append(Paragraph("SCENARIO COMPARISON", ParagraphStyle("SecTitle",
            parent=styles["Normal"], fontSize=11, fontName="Helvetica-Bold",
            textColor=BLUE, spaceAfter=3*mm)))
        headers = ["Scenario","Route","Aircraft","LF","RASK","CASK","Margin","BELF","Daily Profit"]
        sc_data = [headers]
        for sc in scenarios:
            e = sc["economics"]
            sc_data.append([
                sc["name"][:20], f"{e['origin']}→{e['destination']}",
                e["aircraft_type"], f"{e['load_factor']*100:.0f}%",
                f"{e['rask']:.1f}p", f"{e['cask']:.1f}p",
                f"{e['margin_pct']:.1f}%", f"{e['belf']*100:.1f}%",
                f"₹{e['profit_inr']/1e5:.1f}L",
            ])
        sc_style = TableStyle([
            ("BACKGROUND",   (0,0), (-1,0), BLUE),
            ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
            ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 8),
            ("FONTNAME",     (0,1), (-1,-1), "Helvetica"),
            ("TEXTCOLOR",    (0,1), (-1,-1), DGREY),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, GREY]),
            ("GRID",         (0,0), (-1,-1), 0.5, LGREY),
            ("LEFTPADDING",  (0,0), (-1,-1), 6),
            ("RIGHTPADDING", (0,0), (-1,-1), 6),
            ("TOPPADDING",   (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0), (-1,-1), 4),
        ])
        story.append(Table(sc_data, style=sc_style, hAlign="LEFT",
                            colWidths=[W*0.18,W*0.13,W*0.10,W*0.07,
                                       W*0.09,W*0.09,W*0.09,W*0.09,W*0.16]))

    # Footer
    story.append(Spacer(1, 8*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=LGREY))
    story.append(Paragraph(
        "Generated by AeroVia · Route Profitability Simulator · aerovia.streamlit.app",
        ParagraphStyle("Footer", parent=styles["Normal"],
                        fontSize=8, textColor=colors.HexColor("#B0BDD0"),
                        alignment=TA_CENTER, spaceBefore=2*mm)))

    doc.build(story)
    buf.seek(0)
    return buf.read()
